# ==========================================================================
# compile_results.py
# ==========================================================================
# Cross-backbone results compiler for the detection benchmark.
# Source- and stage-agnostic: works for any sensor declared in
# sensor_registry.py (UAV, GE, Aerial, Satellite, or future sources).
#
# Purpose
# -------
# `evaluate_model.py` writes one long-form CSV per (backbone,
# sensor) pair into results/stage_b/ :
#
#     <config_stem>__test_GE.csv
#     <config_stem>__test_aerial.csv
#
# Each such CSV has the schema emitted by metrics_engine.write_csv():
#
#     model , metric , bbox , segm , score_thr , iou_thr , test_set
#
# with one row per metric (12 rows). This script ingests every such CSV,
# transposes each into a single wide row keyed by backbone, and produces
# the compiled cross-backbone tables that the manuscript fill-in requires:
#
#     results/stage_b/compiled/palm_benchmark_test_GE.csv
#     results/stage_b/compiled/palm_benchmark_test_aerial.csv
#     results/stage_b/compiled/palm_benchmark_combined.csv
#     results/stage_b/compiled/palm_benchmark_test_results.xlsx
#
# The XLSX carries one worksheet per sensor plus a Combined sheet, with
# the segm (instance-segmentation) track as the headline metric set and
# the bbox track retained alongside for completeness, consistent with the
# PalmBenchmarkMetric reporting convention.
#
# This script performs NO inference and reads NO dataset paths. It is a
# pure post-processing aggregator over CSVs already on disk. It is
# therefore unaffected by the /workspace/ vs /datasets/ path question:
# that concern belongs solely to evaluate_model.py at evaluation
# time (see the deployment note in the accompanying summary).
#
# --------------------------------------------------------------------------
# Optionally, with --run-missing, this script will invoke
# evaluate_model.py for any backbone whose per-sensor CSV is
# absent, then aggregate. This is a convenience wrapper only; the
# evaluation itself is delegated entirely to the existing Stage B script.
#
# Usage
# -----
#   # Aggregate CSVs already produced by evaluate_model.py:
#   python compile_results.py \
#       --results-dir results/stage_b
#
#   # Aggregate, and first run any backbone whose CSVs are missing:
#   python compile_results.py \
#       --results-dir results/stage_b \
#       --run-missing \
#       --config-dir  configs/Custom/maskrcnn_palm_ms15 \
#       --work-root   work_dirs \
#       --eval-script evaluate_model.py \
#       --backbones   maskrcnn_r50_ms15 maskrcnn_r101_ms15 \
#                     maskrcnn_swin_s_ms15 maskrcnn_pvtv2_b2_ms15 \
#                     maskrcnn_convnext_t_ms15 maskrcnn_mambavision_s_ms15 \
#                     maskrcnn_vmamba_s_ms15 maskrcnn_mambaout_t_ms15
# ==========================================================================

import argparse
import csv
import glob
import os
import os.path as osp
import subprocess
import sys
from collections import OrderedDict
from typing import Dict, List, Optional, Tuple

# Make this folder importable so sensor_registry resolves regardless of
# the working directory the script is launched from.
sys.path.insert(0, osp.dirname(osp.abspath(__file__)))
from sensor_registry import SENSORS, available_sensors  # noqa: E402

# Backbones whose evaluate_model.py subprocess exited non-zero during a
# --run-missing pass. Collected so the stage can finish and report them
# together rather than aborting on the first failure.
_RUN_FAILURES: List[Tuple[str, int]] = []

# --------------------------------------------------------------------------
# 1. CSV schema constants — must mirror metrics_engine.write_csv() exactly
# --------------------------------------------------------------------------
# The per-sensor CSV is long-form: each row carries one metric, with its
# value split across the 'bbox' and 'segm' columns. The 'metric' column
# strings below are the join keys used to locate each value.
#
# (csv_metric_label, short_key) — short_key becomes the wide-form column
# stem; the final column is short_key + '_' + track ('segm' / 'bbox').
METRIC_LABELS: "OrderedDict[str, str]" = OrderedDict([
    ('mAP@50 [PRIMARY]',                    'mAP50'),
    ('F1@0.5 (optimal threshold) [PRIMARY]', 'F1_opt'),
    ('Precision@0.5 (optimal threshold)',    'P_opt'),
    ('Recall@0.5 (optimal threshold)',       'R_opt'),
    ('Best score threshold',                 'best_thr'),
    ('F1@0.5 (fixed score=0.05)',            'F1_fixed'),
    ('Precision@0.5 (fixed score=0.05)',     'P_fixed'),
    ('Recall@0.5 (fixed score=0.05)',        'R_fixed'),
    ('mAP@[.5:.95] [suppl.]',                'mAP'),
    ('TP (fixed score=0.05)',                'TP'),
    ('FP (fixed score=0.05)',                'FP'),
    ('FN (fixed score=0.05)',                'FN'),
])

# Sensor token -> (csv filename suffix, display partition label).
# Built automatically from sensor_registry.py so the compiler covers
# EVERY sensor declared there (GE, Aerial, UAV, Satellite once added)
# with no code change. The 'label' field of each registry entry defines
# both the CSV suffix ('__<label>.csv') and the display label.
SENSOR_SUFFIX: Dict[str, Tuple[str, str]] = {
    key: (f"__{entry['label']}.csv", entry['label'])
    for key, entry in SENSORS.items()
}

# Wide-form column order. The segm track leads (headline); bbox follows.
# Integer-valued columns are listed in INT_COLUMNS for correct formatting.
PRIMARY_COLUMNS: List[str] = [
    'mAP50', 'F1_opt', 'P_opt', 'R_opt', 'best_thr',
    'F1_fixed', 'P_fixed', 'R_fixed', 'mAP', 'TP', 'FP', 'FN',
]
INT_COLUMNS = {'TP', 'FP', 'FN'}

# Efficiency columns (one row per backbone, sensor-independent), produced
# by evaluate_model.write_efficiency_csv() as <config_stem>__efficiency.csv.
# Appended to every sensor row for the matching backbone by left-join.
EFF_SUFFIX = '__efficiency.csv'
EFF_COLUMNS: List[str] = [
    'params_total_M', 'params_trainable_M',
    'gflops', 'flops_uncounted',
    'fps', 'latency_mean_ms', 'latency_std_ms',
    'latency_p50_ms', 'latency_p90_ms',
    'peak_vram_mb',
    'eff_dtype', 'eff_input_h', 'eff_input_w', 'eff_warmup', 'eff_iters',
    'device_name', 'flops_uncounted_ops', 'flops_note', 'flops_scope',
]
EFF_INT_COLUMNS = {'flops_uncounted', 'eff_input_h', 'eff_input_w',
                   'eff_warmup', 'eff_iters'}
EFF_FLOAT_COLUMNS = {'params_total_M', 'params_trainable_M', 'gflops',
                     'fps', 'latency_mean_ms', 'latency_std_ms',
                     'latency_p50_ms', 'latency_p90_ms', 'peak_vram_mb'}


# --------------------------------------------------------------------------
# 2. Long-form CSV parsing
# --------------------------------------------------------------------------
def parse_eval_csv(csv_path: str) -> Optional[dict]:
    """Parses one long-form per-sensor CSV from metrics_engine.write_csv().

    Returns a dict::

        {
          'model'    : <config_stem>,
          'test_set' : 'test_GE' | 'test_aerial',
          'score_thr': <float-or-str>,
          'iou_thr'  : <float-or-str>,
          'segm'     : {short_key: value, ...},
          'bbox'     : {short_key: value, ...},
        }

    Returns None and warns if the file is malformed (missing header,
    unrecognised metric labels), so a single bad CSV does not abort the
    whole compilation.
    """
    with open(csv_path, newline='') as f:
        reader = csv.DictReader(f)
        required = {'model', 'metric', 'bbox', 'segm',
                    'score_thr', 'iou_thr', 'test_set'}
        if reader.fieldnames is None or not required.issubset(
                set(reader.fieldnames)):
            sys.stderr.write(
                f'[WARN] {csv_path}: unexpected header '
                f'{reader.fieldnames}; skipping.\n')
            return None

        rows = list(reader)

    if not rows:
        sys.stderr.write(f'[WARN] {csv_path}: empty file; skipping.\n')
        return None

    out = {
        'model': rows[0]['model'],
        'test_set': rows[0]['test_set'],
        'score_thr': rows[0]['score_thr'],
        'iou_thr': rows[0]['iou_thr'],
        'segm': {},
        'bbox': {},
    }

    label_to_key = dict(METRIC_LABELS)
    seen = set()
    for row in rows:
        label = row['metric']
        key = label_to_key.get(label)
        if key is None:
            # A metric label not in the known set: tolerate but note it.
            sys.stderr.write(
                f'[WARN] {csv_path}: unrecognised metric label '
                f'"{label}"; ignoring this row.\n')
            continue
        out['segm'][key] = _to_number(row['segm'])
        out['bbox'][key] = _to_number(row['bbox'])
        seen.add(key)

    missing = set(METRIC_LABELS.values()) - seen
    if missing:
        sys.stderr.write(
            f'[WARN] {csv_path}: missing metric rows {sorted(missing)}; '
            f'those cells will be blank.\n')
    return out


def _to_number(text: str):
    """Casts a CSV cell to int or float where possible; '' -> None."""
    if text is None or text == '':
        return None
    try:
        f = float(text)
        return int(f) if f.is_integer() else f
    except ValueError:
        return text


# --------------------------------------------------------------------------
# 3. Backbone-label normalisation
# --------------------------------------------------------------------------
def tidy_backbone(config_stem: str) -> str:
    """Derives a compact, stage-independent backbone label from a config stem.

    The benchmark spans four stages whose configs carry different suffixes:
        Stage A : maskrcnn_<backbone>_uav5cm
        Stage B : maskrcnn_<backbone>_ms15
        Stage C : maskrcnn_<backbone>_stagec
        Stage D : maskrcnn_<backbone>_sat  /  maskrcnn_<backbone>_staged_ft
    Stripping only one suffix (as an earlier version did) left the stage
    token attached for every stage except B, so the same architecture
    appeared under different keys across stages and the cross-stage tables
    could not be joined. This routine removes the leading 'maskrcnn_' and
    any one recognised trailing stage suffix, yielding a clean architecture
    label (e.g. 'r50', 'mambavision_s', 'spatialmamba_s') identical across
    all stages. Unrecognised stems are returned unchanged so nothing is
    silently lost.
    """
    stem = config_stem
    if stem.startswith('maskrcnn_'):
        stem = stem[len('maskrcnn_'):]
    # Longest suffixes first, so '_staged_ft' is matched before '_sat'-like
    # fragments and '_uav5cm' is not partially consumed.
    for suffix in ('_staged_ft', '_uav5cm', '_stagec', '_ms15', '_sat',
                   '_staged'):
        if stem.endswith(suffix):
            stem = stem[:-len(suffix)]
            break
    return stem or config_stem


# --------------------------------------------------------------------------
# 4. Optional: run evaluate_model.py for missing backbones
# --------------------------------------------------------------------------
def _iter_number(ckpt_path: str) -> int:
    """Extracts the trailing iteration number from a checkpoint filename.

    Handles 'best_<metric>_iter_25000.pth', 'iter_25000.pth', and
    'epoch_36.pth'. Returns -1 if no number is found, so such files sort
    last and are never preferred over a numbered one.
    """
    import re
    m = re.findall(r'(\d+)', osp.splitext(osp.basename(ckpt_path))[0])
    return int(m[-1]) if m else -1


def find_checkpoint(work_root: str, config_stem: str,
                    ckpt_pattern: Optional[str] = None) -> Optional[str]:
    """Locates the checkpoint to evaluate inside one model's work_dir.

    Source-agnostic by design. The CheckpointHook 'save_best' filename
    depends on the monitored metric key, which differs across stages:
        Stage B  : best_coco_segm_mAP_50_iter_*.pth
        Stage C  : best_UAV_segm_mAP_50_iter_*.pth,
                   best_GE_segm_mAP_50_iter_*.pth   (MultiDatasetsEvaluator)
        future   : any best_*.pth
    Rather than hardcode one pattern, this function matches ANY
    'best_*.pth'. Resolution order:

      1. If ckpt_pattern is given, use it verbatim as the glob (relative
         to work_dir). This is the unambiguous selector when a Stage C
         work_dir contains two per-sensor best checkpoints and a
         specific one is wanted, e.g. 'best_GE_*'.
      2. Otherwise, glob 'best_*.pth'. If exactly one matches, use it.
         If several match (e.g. Stage C dual-sensor), the one with the
         highest iteration number is chosen and a warning lists the
         alternatives so the choice is visible and overridable.
      3. Fallback: highest-iteration 'iter_*.pth' / 'epoch_*.pth'.

    Returns the checkpoint path, or None (with a warning) if none found.
    """
    work_dir = osp.join(work_root, config_stem)
    if not osp.isdir(work_dir):
        sys.stderr.write(f'[WARN] work_dir not found: {work_dir}\n')
        return None

    # 1. Explicit user-supplied pattern.
    if ckpt_pattern:
        matches = sorted(glob.glob(osp.join(work_dir, ckpt_pattern)),
                         key=_iter_number)
        if matches:
            if len(matches) > 1:
                sys.stderr.write(
                    f'[WARN] {config_stem}: pattern "{ckpt_pattern}" '
                    f'matched {len(matches)} files; using highest-iter '
                    f'{osp.basename(matches[-1])}.\n')
            return matches[-1]
        sys.stderr.write(
            f'[WARN] {config_stem}: pattern "{ckpt_pattern}" matched '
            f'nothing in {work_dir}; falling back to best_*.pth.\n')

    # 2. Any best_*.pth checkpoint.
    best = sorted(glob.glob(osp.join(work_dir, 'best_*.pth')),
                  key=_iter_number)
    if best:
        if len(best) > 1:
            names = ', '.join(osp.basename(p) for p in best)
            sys.stderr.write(
                f'[WARN] {config_stem}: multiple best_* checkpoints '
                f'present [{names}]; using highest-iter '
                f'{osp.basename(best[-1])}. Pass --ckpt-pattern to '
                f'select a specific one (e.g. "best_GE_*").\n')
        return best[-1]

    # 3. Fallback: highest-iteration iter_*/epoch_* checkpoint.
    iters = sorted(
        glob.glob(osp.join(work_dir, 'iter_*.pth'))
        + glob.glob(osp.join(work_dir, 'epoch_*.pth')),
        key=_iter_number,
    )
    if iters:
        sys.stderr.write(
            f'[WARN] {config_stem}: no best_* checkpoint; using '
            f'{osp.basename(iters[-1])}.\n')
        return iters[-1]

    sys.stderr.write(f'[WARN] no checkpoint in {work_dir}\n')
    return None


def run_missing(args) -> None:
    """Invokes evaluate_model.py for any backbone lacking CSVs."""
    for config_stem in args.backbones:
        cfg_path = osp.join(args.config_dir, config_stem + '.py')
        if not osp.isfile(cfg_path):
            sys.stderr.write(
                f'[WARN] config missing, cannot evaluate: {cfg_path}\n')
            continue

        # Determine which sensors still need a CSV. When --sensors is
        # given, only those keys are considered; otherwise every sensor
        # declared in sensor_registry.py is.
        target_sensors = (args.sensors if args.sensors
                          else list(SENSOR_SUFFIX.keys()))
        pending = []
        for sensor in target_sensors:
            suffix = SENSOR_SUFFIX[sensor][0]
            csv_path = osp.join(args.results_dir, config_stem + suffix)
            if not osp.isfile(csv_path):
                pending.append(sensor)
        if not pending:
            print(f'[skip] {config_stem}: CSVs for {target_sensors} '
                  f'already present.')
            continue

        ckpt = find_checkpoint(args.work_root, config_stem,
                               ckpt_pattern=args.ckpt_pattern)
        if ckpt is None:
            sys.stderr.write(
                f'[WARN] {config_stem}: no checkpoint; cannot evaluate.\n')
            continue

        cmd = [
            sys.executable, args.eval_script,
            '--config', cfg_path,
            '--checkpoint', ckpt,
            '--sensors', *pending,
            '--results-dir', args.results_dir,
        ]
        if args.batch_size is not None:
            cmd += ['--batch-size', str(args.batch_size)]
        if args.num_workers is not None:
            cmd += ['--num-workers', str(args.num_workers)]
        if args.device is not None:
            cmd += ['--device', str(args.device)]
        if getattr(args, 'max_dets', None) is not None:
            cmd += ['--max-dets', str(args.max_dets)]
        if getattr(args, 'applied_score_thr', None) is not None:
            cmd += ['--applied-score-thr', str(args.applied_score_thr)]
        print(f'\n[run ] {config_stem}: evaluating sensors {pending}')
        print('       $ ' + ' '.join(cmd))
        result = subprocess.run(cmd, check=False)
        if result.returncode != 0:
            # A single backbone failing (tracer abort, OOM, missing kernel,
            # ...) must not abort the whole stage. Any accuracy CSV already
            # written before the failure is retained and will be picked up
            # by the aggregator; the backbone is recorded and the run
            # continues. Re-run that backbone individually to diagnose
            # (e.g. with --no-flops if the failure was in the FLOP trace).
            _RUN_FAILURES.append((config_stem, result.returncode))
            sys.stderr.write(
                f'[WARN] {config_stem}: evaluate_model.py exited with code '
                f'{result.returncode}; continuing with remaining backbones. '
                f'Any accuracy CSV written before the failure is kept.\n')

    if _RUN_FAILURES:
        print('\n' + '=' * 72)
        print(f'  {len(_RUN_FAILURES)} backbone(s) failed during --run-missing:')
        for stem, code in _RUN_FAILURES:
            print(f'    {stem}  (exit {code})')
        print('  Their accuracy CSVs (if reached) are still aggregated below.')
        print('  Re-run a failed backbone alone to diagnose; a FLOP-trace '
              'abort\n  clears with --no-flops.')
        print('=' * 72)


# --------------------------------------------------------------------------
# 5. Aggregation
# --------------------------------------------------------------------------
def collect_records(results_dir: str) -> Dict[str, List[dict]]:
    """Scans results_dir for per-sensor CSVs and groups parsed records
    by sensor token. Returns {'GE': [rec, ...], 'Aerial': [rec, ...]}.
    """
    grouped: Dict[str, List[dict]] = {s: [] for s in SENSOR_SUFFIX}
    for sensor, (suffix, _) in SENSOR_SUFFIX.items():
        pattern = osp.join(results_dir, f'*{suffix}')
        for csv_path in sorted(glob.glob(pattern)):
            rec = parse_eval_csv(csv_path)
            if rec is None:
                continue
            rec['backbone'] = tidy_backbone(rec['model'])
            rec['source_csv'] = osp.basename(csv_path)
            grouped[sensor].append(rec)
    return grouped


def collect_efficiency(results_dir: str) -> Dict[str, dict]:
    """Loads every <config_stem>__efficiency.csv into a per-backbone map.

    Returns {backbone_label: {eff_col: value, ...}}. The efficiency CSV is
    keyed by config_stem; tidy_backbone() maps that to the same compact
    label used by the accuracy wide rows, so the subsequent join is exact.
    String columns (device_name, dtype, uncounted-op list, note) are kept
    verbatim; numeric columns are cast for correct downstream formatting.
    """
    eff: Dict[str, dict] = {}
    for csv_path in sorted(glob.glob(osp.join(results_dir, f'*{EFF_SUFFIX}'))):
        with open(csv_path, newline='') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        if not rows:
            sys.stderr.write(f'[WARN] {csv_path}: empty efficiency CSV.\n')
            continue
        rec = rows[0]
        model = rec.get('model', '')
        backbone = tidy_backbone(model)
        parsed: dict = {}
        for col in EFF_COLUMNS:
            raw = rec.get(col, '')
            if col in EFF_FLOAT_COLUMNS or col in EFF_INT_COLUMNS:
                parsed[col] = _to_number(raw)
            else:
                parsed[col] = (raw if raw != '' else None)
        eff[backbone] = parsed
    return eff



def build_wide_rows(records: List[dict],
                    efficiency: Optional[Dict[str, dict]] = None
                    ) -> List[OrderedDict]:
    """Transposes long-form records into wide rows, one per backbone.

    Column layout: Backbone, Partition, then segm_<metric> for every
    PRIMARY_COLUMNS entry, then bbox_<metric> for the same set, then the
    efficiency columns (left-joined by backbone from the efficiency map,
    if provided), then provenance columns (score_thr, iou_thr, Source CSV).

    Efficiency is sensor-independent: the same per-backbone figures are
    attached to every sensor row for that backbone. A backbone with no
    efficiency CSV gets blank efficiency cells rather than being dropped.
    """
    efficiency = efficiency or {}
    rows: List[OrderedDict] = []
    for rec in sorted(records, key=lambda r: r['backbone']):
        row: "OrderedDict[str, object]" = OrderedDict()
        row['Backbone'] = rec['backbone']
        row['Partition'] = rec['test_set']
        for track in ('segm', 'bbox'):
            for key in PRIMARY_COLUMNS:
                row[f'{track}_{key}'] = rec[track].get(key)
        eff_rec = efficiency.get(rec['backbone'], {})
        for col in EFF_COLUMNS:
            # Some efficiency columns already carry an 'eff_' prefix
            # (eff_dtype, eff_input_h, ...); avoid doubling it.
            out_col = col if col.startswith('eff_') else f'eff_{col}'
            row[out_col] = eff_rec.get(col)
        row['score_thr'] = rec['score_thr']
        row['iou_thr'] = rec['iou_thr']
        row['Source CSV'] = rec['source_csv']
        rows.append(row)
    return rows


# --------------------------------------------------------------------------
# 6. CSV output
# --------------------------------------------------------------------------
def write_wide_csv(rows: List[OrderedDict], out_path: str,
                   warn_if_empty: bool = True) -> None:
    if not rows:
        if warn_if_empty:
            sys.stderr.write(f'[WARN] no rows to write for {out_path}\n')
        return
    os.makedirs(osp.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in rows:
            w.writerow({k: ('' if v is None else v) for k, v in r.items()})
    print(f'  CSV  -> {out_path}')


# --------------------------------------------------------------------------
# 7. XLSX output
# --------------------------------------------------------------------------
def write_xlsx(sensor_rows: Dict[str, List[OrderedDict]],
               combined_rows: List[OrderedDict],
               out_path: str) -> None:
    """Writes a formatted workbook: one sheet per sensor + Combined.

    The F1 columns (segm_F1_opt, segm_F1_fixed, bbox_F1_opt,
    bbox_F1_fixed) are written as live Excel formulas referencing the
    corresponding precision and recall cells, so the workbook remains
    auditable and recomputes if a precision/recall cell is edited.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.stderr.write(
            '[WARN] openpyxl not installed; XLSX skipped. '
            'Install with: pip install openpyxl\n')
        return

    header_fill = PatternFill('solid', start_color='1F4E78')
    segm_fill   = PatternFill('solid', start_color='2E75B6')
    bbox_fill   = PatternFill('solid', start_color='808080')
    eff_fill    = PatternFill('solid', start_color='548235')  # efficiency
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    body_font   = Font(name='Arial', size=10)

    # F1 column -> (precision column, recall column) for formula wiring.
    f1_pairs = {
        'segm_F1_opt':   ('segm_P_opt',   'segm_R_opt'),
        'segm_F1_fixed': ('segm_P_fixed', 'segm_R_fixed'),
        'bbox_F1_opt':   ('bbox_P_opt',   'bbox_R_opt'),
        'bbox_F1_fixed': ('bbox_P_fixed', 'bbox_R_fixed'),
    }

    def _sheet(ws, rows: List[OrderedDict]):
        if not rows:
            ws.append(['(no data)'])
            return
        headers = list(rows[0].keys())
        col_index = {h: i + 1 for i, h in enumerate(headers)}  # 1-based

        ws.append(headers)
        for h in headers:
            cell = ws.cell(row=1, column=col_index[h])
            cell.font = header_font
            if h.startswith('segm_'):
                cell.fill = segm_fill
            elif h.startswith('bbox_'):
                cell.fill = bbox_fill
            elif h.startswith('eff_'):
                cell.fill = eff_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center', wrap_text=True)

        for i, r in enumerate(rows):
            excel_row = i + 2
            ws.append([('' if r[h] is None else r[h]) for h in headers])
            # Overwrite F1 cells with live formulas.
            for f1_col, (p_col, r_col) in f1_pairs.items():
                if f1_col in col_index and p_col in col_index \
                        and r_col in col_index:
                    pL = get_column_letter(col_index[p_col])
                    rL = get_column_letter(col_index[r_col])
                    ws.cell(row=excel_row, column=col_index[f1_col]).value = (
                        f'=IFERROR(2*{pL}{excel_row}*{rL}{excel_row}/'
                        f'({pL}{excel_row}+{rL}{excel_row}),0)')
            for h in headers:
                ws.cell(row=excel_row, column=col_index[h]).font = body_font
                ws.cell(row=excel_row, column=col_index[h]).alignment = \
                    Alignment(horizontal='center')

        # Number formats.
        def _effcol(c):
            return c if c.startswith('eff_') else f'eff_{c}'
        eff_int = {_effcol(c) for c in EFF_INT_COLUMNS}
        eff_float4 = {'eff_params_total_M', 'eff_params_trainable_M',
                      'eff_gflops', 'eff_fps'}
        eff_float2 = {'eff_latency_mean_ms', 'eff_latency_std_ms',
                      'eff_latency_p50_ms', 'eff_latency_p90_ms',
                      'eff_peak_vram_mb'}
        for h in headers:
            stem = h.split('_', 1)[-1] if '_' in h else h
            col = col_index[h]
            for excel_row in range(2, len(rows) + 2):
                cell = ws.cell(row=excel_row, column=col)
                if h in eff_int:
                    cell.number_format = '#,##0'
                elif h in eff_float4:
                    cell.number_format = '0.0000'
                elif h in eff_float2:
                    cell.number_format = '0.00'
                elif stem in INT_COLUMNS:
                    cell.number_format = '#,##0'
                elif stem == 'best_thr':
                    cell.number_format = '0.000'
                elif h.startswith(('segm_', 'bbox_')):
                    cell.number_format = '0.0000'

        # Column widths.
        for h in headers:
            width = 30 if h == 'Source CSV' else (
                15 if h in ('Backbone', 'Partition') else 12)
            ws.column_dimensions[
                get_column_letter(col_index[h])].width = width
        ws.freeze_panes = 'C2'   # keep Backbone + Partition visible

    wb = Workbook()
    wb.remove(wb.active)
    for sensor, (_, label) in SENSOR_SUFFIX.items():
        _sheet(wb.create_sheet(label), sensor_rows.get(sensor, []))
    _sheet(wb.create_sheet('Combined'), combined_rows)

    os.makedirs(osp.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'  XLSX -> {out_path}')


# --------------------------------------------------------------------------
# 8. Console summary
# --------------------------------------------------------------------------
def print_summary(sensor_rows: Dict[str, List[OrderedDict]]) -> None:
    line = '=' * 84
    print(f'\n{line}')
    print('  COMPILED STAGE B TEST RESULTS — segm track (PRIMARY)')
    print(line)
    hdr = f'  {"Backbone":<18}{"Partition":<14}' \
          f'{"mAP@50":>10}{"F1_opt":>10}{"P_opt":>10}{"R_opt":>10}'
    for sensor, (_, label) in SENSOR_SUFFIX.items():
        rows = sensor_rows.get(sensor, [])
        if not rows:
            continue
        print(hdr)
        print('  ' + '-' * 80)
        for r in rows:
            def g(k):
                v = r.get(k)
                return f'{v:.4f}' if isinstance(v, float) else (
                    '' if v is None else str(v))
            print(f'  {r["Backbone"]:<18}{r["Partition"]:<14}'
                  f'{g("segm_mAP50"):>10}{g("segm_F1_opt"):>10}'
                  f'{g("segm_P_opt"):>10}{g("segm_R_opt"):>10}')
        print()
    print(line)


# --------------------------------------------------------------------------
# 9. CLI
# --------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description='Compile per-(backbone, sensor) Stage B evaluation '
                    'CSVs into cross-backbone GE/Aerial fill-in tables.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument('--results-dir', default='results/stage_b',
                   help='Directory holding the per-sensor CSVs written by '
                        'evaluate_model.py. Default: results/stage_b')
    p.add_argument('--out-dir', default=None,
                   help='Output directory for the compiled tables. '
                        'Default: <results-dir>/compiled')
    # --run-missing optional evaluation driver.
    p.add_argument('--sensors', nargs='+', default=None, metavar='SENSOR',
                   help='Restrict evaluation/aggregation to these sensor '
                        'keys (from sensor_registry.py). Default: all '
                        'declared sensors.')
    p.add_argument('--ckpt-pattern', default=None, metavar='GLOB',
                   help='Glob (relative to each work_dir) selecting which '
                        'checkpoint to evaluate, e.g. "best_GE_*". Use when '
                        'a work_dir holds several best_* checkpoints. '
                        'Default: highest-iteration best_*.pth.')
    p.add_argument('--run-missing', action='store_true',
                   help='Before aggregating, invoke evaluate_model.py '
                        'for any backbone whose per-sensor CSV is absent.')
    p.add_argument('--eval-script', default='evaluate_model.py',
                   help='Path to evaluate_model.py '
                        '(used only with --run-missing).')
    p.add_argument('--config-dir', default=None,
                   help='Directory of maskrcnn_<backbone>_ms15.py configs '
                        '(used only with --run-missing).')
    p.add_argument('--work-root', default=None,
                   help='Root of per-backbone training work_dirs '
                        '(used only with --run-missing).')
    p.add_argument('--backbones', nargs='+', default=None,
                   help='Config stems to ensure are evaluated, e.g. '
                        'maskrcnn_r50_ms15 (used only with --run-missing).')
    p.add_argument('--batch-size', type=int, default=None,
                   help='Forwarded to evaluate_model.py.')
    p.add_argument('--num-workers', type=int, default=None,
                   help='Forwarded to evaluate_model.py.')
    p.add_argument('--max-dets', type=int, default=None,
                   help='Per-image detection cap, forwarded to '
                        'evaluate_model.py for any backbone re-run under '
                        '--run-missing. Set to the benchmark-wide value '
                        '(e.g. 500) so every stage and sensor is scored '
                        'under one cap.')
    p.add_argument('--applied-score-thr', type=float, default=None,
                   help='Fixed (validation-selected) operating threshold, '
                        'forwarded to evaluate_model.py under '
                        '--run-missing.')
    p.add_argument('--device', default='auto', metavar='DEVICE',
                   help='Compute device, forwarded to '
                        'evaluate_model.py. "auto" (default) uses '
                        'the GPU and aborts if none is found. See that '
                        'script for "cuda"/"cuda:N"/"cpu" options.')
    return p.parse_args()


def main() -> None:
    args = parse_args()
    out_dir = args.out_dir or osp.join(args.results_dir, 'compiled')

    # Validate requested sensor keys against the registry up front, so an
    # unknown key fails with a clear message instead of a KeyError later.
    if args.sensors:
        valid = available_sensors()
        unknown = [s for s in args.sensors if s not in valid]
        if unknown:
            sys.exit(f'[ERROR] unknown sensor(s) {unknown}. Valid keys in '
                     f'sensor_registry.py: {valid}')

    if args.run_missing:
        missing_req = [a for a in ('config_dir', 'work_root', 'backbones')
                       if getattr(args, a) is None]
        if missing_req:
            sys.exit(f'[ERROR] --run-missing requires '
                     f'{["--" + m.replace("_", "-") for m in missing_req]}')
        run_missing(args)

    if not osp.isdir(args.results_dir):
        sys.exit(f'[ERROR] results-dir not found: {args.results_dir}')

    grouped = collect_records(args.results_dir)
    efficiency = collect_efficiency(args.results_dir)
    sensor_rows = {s: build_wide_rows(recs, efficiency)
                   for s, recs in grouped.items()}

    n_total = sum(len(v) for v in sensor_rows.values())
    if n_total == 0:
        expected = ' / '.join(
            f'*{suffix}' for suffix, _ in SENSOR_SUFFIX.values())
        sys.exit(f'[ERROR] no per-sensor CSVs found in {args.results_dir}. '
                 f'Expected files matching: {expected}')

    # Per-sensor wide CSVs. A compiled CSV is written only for sensors
    # that actually produced rows; sensors declared in the registry but
    # not part of this run are skipped silently (no spurious warning).
    for sensor, (_, label) in SENSOR_SUFFIX.items():
        rows = sensor_rows.get(sensor, [])
        write_wide_csv(rows,
                       osp.join(out_dir, f'palm_benchmark_{label}.csv'),
                       warn_if_empty=bool(rows))

    # Combined CSV: every sensor's rows concatenated, in registry order.
    # Built from whatever sensors are present, so Stage A (UAV-only),
    # Stage D (Sat-only), or any subset works without a KeyError.
    combined: List[OrderedDict] = []
    for sensor in SENSOR_SUFFIX:
        combined.extend(sensor_rows.get(sensor, []))
    write_wide_csv(combined,
                   osp.join(out_dir, 'palm_benchmark_combined.csv'))

    # XLSX workbook.
    write_xlsx(sensor_rows, combined,
               osp.join(out_dir, 'palm_benchmark_test_results.xlsx'))

    print_summary(sensor_rows)
    print(f'{n_total} (backbone, sensor) record(s) compiled into {out_dir}')


if __name__ == '__main__':
    main()
